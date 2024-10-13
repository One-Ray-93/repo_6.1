
from openpyxl.reader.excel import load_workbook


workbook = load_workbook("TMP/File_1.xlsx")
sheet = workbook.active

print(sheet.cell(row=1, column=2).value
)
print(len(sheet.columns)

)
